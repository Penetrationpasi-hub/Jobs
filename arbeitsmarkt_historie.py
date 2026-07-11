#!/usr/bin/env python3
"""
Historische Arbeitsmarktdaten mit Zukunftsperspektive.

Zieht lange Zeitreihen aus der Regionaldatenbank Deutschland (GENESIS)
fuer Rems-Murr-Kreis, Ostalbkreis und Stadtkreis Stuttgart und baut
daraus arbeitsmarkt_historie.xlsx mit Rueckblick (ab 2008) und einem
vorsichtigen Ausblick (Demografie, Trendfortschreibung als Illustration).

Datenquellen (EVAS):
  13211  Arbeitslose je Kreis, Monatswerte
  13111  Sozialversicherungspflichtig Beschaeftigte je Kreis, Stichtage
  12411  Bevoelkerung nach Altersgruppen je Kreis (Fortschreibung)
  12421  Regionalisierte Bevoelkerungsvorausberechnung je Kreis

Zugang: kostenloses Konto auf regionalstatistik.de,
env GENESIS_USER / GENESIS_PASS (gleiche Secrets wie collect_jobs.py).
Tabellencodes lassen sich per env ueberschreiben (TAB_ALO, TAB_BESCH,
TAB_BEV, TAB_PROG). Bei Fehlern listet das Script den Katalog.

Outputs:
  data_historie/*.csv           eine Datei je Statistik (Langformat)
  arbeitsmarkt_historie.xlsx    Auswertung mit Diagrammen
"""

import csv
import json
import os
import sys
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import AreaChart3D, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
DATA = BASE / "data_historie"
OUT = BASE / "arbeitsmarkt_historie.xlsx"

GENESIS_BASE = "https://www.regionalstatistik.de/genesisWS/rest/2020"
STARTYEAR = int(os.environ.get("STARTYEAR", "2008"))
USER_AGENT = "arbeitsmarkt-historie/1.0 (privates Arbeitsmarkt-Monitoring)"

KREISE = {
    "08111": "Stuttgart",
    "08119": "Rems-Murr-Kreis",
    "08136": "Ostalbkreis",
}

# key -> (Tabellencode, Beschreibung, Merkmals-Praefixe die als
# Dimension erhalten bleiben statt auf "Insgesamt" gefiltert zu werden)
TABELLEN = {
    "arbeitslose": (
        os.environ.get("TAB_ALO", "13211-01-03-4-B"),
        "Arbeitslose je Kreis, Monatswerte", ()),
    "beschaeftigte": (
        os.environ.get("TAB_BESCH", "13111-01-03-4"),
        "SV-Beschaeftigte je Kreis, Stichtage", ()),
    "bevoelkerung": (
        os.environ.get("TAB_BEV", "12411-04-02-4"),
        "Bevoelkerung nach Altersgruppen je Kreis", ("ALT",)),
    "prognose": (
        os.environ.get("TAB_PROG", "12421-01-01-4"),
        "Bevoelkerungsvorausberechnung je Kreis", ("ALT", "VAR", "BEVP")),
}

ARIAL = "Arial"
HEAD_FONT = Font(name=ARIAL, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", start_color="1D3C6E")
MONAT_NUM = {f"MONAT{m:02d}": m for m in range(1, 13)}


# ---------------- GENESIS-Zugriff -------------------------------------

def creds():
    user = os.environ.get("GENESIS_USER", "").strip()
    pw = os.environ.get("GENESIS_PASS", "").strip()
    if not user or not pw:
        return None
    return user, pw


def genesis_post(endpoint: str, params: dict) -> bytes:
    user, pw = creds()
    body = urllib.parse.urlencode(params).encode("utf-8")
    req = urllib.request.Request(
        f"{GENESIS_BASE}/{endpoint}", data=body, method="POST",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "username": user,
            "password": pw,
            "User-Agent": USER_AGENT,
        })
    with urllib.request.urlopen(req, timeout=300) as resp:
        return resp.read()


def katalog_hinweis(muster: str):
    """Verfuegbare Tabellen zu einem EVAS-Muster listen (Diagnose)."""
    try:
        raw = genesis_post("catalogue/tables", {
            "selection": muster, "area": "free",
            "pagelength": 80, "language": "de",
        })
        data = json.loads(raw.decode("utf-8", errors="replace"))
        for t in (data.get("List") or [])[:80]:
            print(f"  Tabelle: {t.get('Code')} - {t.get('Content')}",
                  file=sys.stderr)
        print("Passenden Code per env TAB_... setzen.", file=sys.stderr)
    except Exception as e:
        print(f"Katalogabruf fehlgeschlagen: {e}", file=sys.stderr)


# ---------------- ffcsv-Parser ----------------------------------------

def parse_zahl(s):
    s = str(s or "").strip()
    if s in ("", "-", ".", "...", "x", "/"):
        return None
    try:
        f = float(s.replace(",", "."))
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


def jahr_aus_zeit(s: str):
    """Zeit-Spalte kann Jahr (2020) oder Stichtag (31.12.2020) sein."""
    s = (s or "").strip()
    teil = s.split(".")[-1]
    try:
        return int(teil)
    except ValueError:
        return None


def parse_ffcsv(text: str, keep_praefixe: tuple):
    """Flat-File-CSV generisch in Langformat-Zeilen zerlegen.

    Ergebnis: Liste von Dicts mit jahr, monat, ags, kreis, dim,
    merkmal (Label der Wertspalte), wert.
    Zeilen mit weiteren Dimensionen werden nur behalten, wenn deren
    Auspraegung "Insgesamt" ist oder der Merkmal-Code mit einem der
    keep_praefixe beginnt (dann landet das Label in "dim").
    """
    reader = csv.DictReader(text.splitlines(), delimiter=";")
    felder = reader.fieldnames or []
    wertspalten = [f for f in felder if "__" in f]
    # Dimensionspaare (n_Merkmal_Code, n_Auspraegung_Code/Label) einsammeln
    dims = []
    n = 1
    while f"{n}_Merkmal_Code" in felder:
        dims.append(n)
        n += 1
    zeilen = []
    for row in reader:
        jahr = jahr_aus_zeit(row.get("Zeit", ""))
        if jahr is None:
            continue
        monat = None
        ags = None
        extra = []
        ok = True
        for d in dims:
            mcode = (row.get(f"{d}_Merkmal_Code") or "").strip()
            acode = (row.get(f"{d}_Auspraegung_Code") or "").strip()
            alabel = (row.get(f"{d}_Auspraegung_Label") or "").strip()
            if acode in KREISE:
                ags = acode
            elif mcode == "MONAT" or acode in MONAT_NUM:
                monat = MONAT_NUM.get(acode)
            elif alabel == "Insgesamt" or not alabel:
                continue
            elif any(mcode.startswith(p) for p in keep_praefixe):
                extra.append(alabel)
            else:
                ok = False
                break
        if not ok or ags is None:
            continue
        for ws in wertspalten:
            wert = parse_zahl(row.get(ws, ""))
            if wert is None:
                continue
            teile = ws.split("__")
            label = teile[1].replace("_", " ") if len(teile) > 1 else ws
            zeilen.append({
                "jahr": jahr, "monat": monat, "ags": ags,
                "kreis": KREISE[ags],
                "dim": " | ".join(extra) if extra else "Insgesamt",
                "merkmal": label, "wert": wert,
            })
    return zeilen


# ---------------- Abruf und Ablage ------------------------------------

def fetch_all():
    if creds() is None:
        print("GENESIS_USER/GENESIS_PASS nicht gesetzt, Abruf "
              "uebersprungen, Excel nutzt vorhandene CSVs.",
              file=sys.stderr)
        return
    DATA.mkdir(exist_ok=True)
    jahr = date.today().year
    for key, (code, beschr, keep) in TABELLEN.items():
        params = {
            "name": code,
            "area": "free",
            "regionalvariable": "KREISE",
            "regionalkey": ",".join(KREISE),
            "startyear": STARTYEAR,
            "endyear": jahr + 25,      # Vorausberechnung reicht in die Zukunft
            "format": "ffcsv",
            "language": "de",
            "compress": "false",
        }
        print(f"Hole {key} ({code}) ...")
        try:
            raw = genesis_post("data/tablefile", params)
        except Exception as e:
            print(f"{key}: Abruf fehlgeschlagen: {e}", file=sys.stderr)
            katalog_hinweis(code.split("-")[0] + "*")
            continue
        text = raw.decode("utf-8-sig", errors="replace")
        if text.lstrip().startswith("{"):
            print(f"{key}: GENESIS-Fehlerantwort: {text[:300]}",
                  file=sys.stderr)
            katalog_hinweis(code.split("-")[0] + "*")
            continue
        zeilen = parse_ffcsv(text, keep)
        if not zeilen:
            print(f"{key}: keine verwertbaren Zeilen.", file=sys.stderr)
            continue
        pfad = DATA / f"{key}.csv"
        with pfad.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["jahr", "monat", "ags", "kreis",
                        "dim", "merkmal", "wert"])
            for z in sorted(zeilen, key=lambda z: (
                    z["jahr"], z["monat"] or 0, z["ags"],
                    z["dim"], z["merkmal"])):
                w.writerow([z["jahr"], z["monat"] or "", z["ags"],
                            z["kreis"], z["dim"], z["merkmal"], z["wert"]])
        print(f"{key}: {len(zeilen)} Zeilen -> {pfad.name}")


def lade(key):
    pfad = DATA / f"{key}.csv"
    if not pfad.exists():
        return []
    rows = []
    with pfad.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append({
                "jahr": int(row["jahr"]),
                "monat": int(row["monat"]) if row["monat"] else None,
                "ags": row["ags"], "kreis": row["kreis"],
                "dim": row["dim"], "merkmal": row["merkmal"],
                "wert": parse_zahl(row["wert"]),
            })
    return rows


# ---------------- Hilfen fuer die Auswertung --------------------------

def merkmal_wahl(rows, *begriffe):
    """Erstes Merkmal-Label finden, das alle Begriffe enthaelt."""
    for r in rows:
        low = r["merkmal"].lower()
        if all(b.lower() in low for b in begriffe):
            return r["merkmal"]
    return rows[0]["merkmal"] if rows else None


def alter_grenzen(label: str):
    """Altersgruppen-Label wie '15 bis unter 25 Jahre' -> (15, 25)."""
    import re
    low = label.lower()
    zahlen = [int(z) for z in re.findall(r"\d+", low)]
    if "unter" in low and len(zahlen) == 1:
        return (0, zahlen[0])
    if ("und mehr" in low or "und aelter" in low
            or "und älter" in low) and len(zahlen) == 1:
        return (zahlen[0], 200)
    if len(zahlen) >= 2:
        return (zahlen[0], zahlen[1])
    return None


def summe_altersband(rows, jahr, ags, merkmal, von, bis):
    """Bevoelkerung im Band [von, bis) aus Altersgruppen aufsummieren."""
    s = 0
    tref = False
    for r in rows:
        if (r["jahr"] != jahr or r["ags"] != ags
                or r["merkmal"] != merkmal or r["dim"] == "Insgesamt"):
            continue
        g = alter_grenzen(r["dim"])
        if g and g[0] >= von and g[1] <= bis and r["wert"] is not None:
            s += r["wert"]
            tref = True
    return s if tref else None


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        z = ws.cell(row=row, column=c)
        z.font = HEAD_FONT
        z.fill = HEAD_FILL


def flaeche3d(titel, hoehe=9, breite=18):
    ch = AreaChart3D()
    ch.grouping = "standard"
    ch.title = titel
    ch.height, ch.width = hoehe, breite
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.z_axis.delete = False
    ch.y_axis.number_format = "#,##0"
    ch.legend.position = "b"
    return ch


# ---------------- Excel -----------------------------------------------

def build_excel():
    alo = lade("arbeitslose")
    bes = lade("beschaeftigte")
    bev = lade("bevoelkerung")
    prog = lade("prognose")
    ags_liste = list(KREISE)
    heute = date.today()

    wb = Workbook()
    zs = wb.active
    zs.title = "Zusammenfassung"
    zs["A1"] = "Historische Arbeitsmarktdaten mit Zukunftsperspektive"
    zs["A1"].font = Font(name=ARIAL, bold=True, size=14)
    zs["A2"] = "Stand"
    zs["B2"] = heute
    zs["B2"].number_format = "DD.MM.YYYY"
    zs["A3"] = "Kreise"
    zs["B3"] = ", ".join(KREISE.values())
    zs["A4"] = "Quelle"
    zs["B4"] = ("Regionaldatenbank Deutschland (GENESIS), "
                "Daten der Statistischen Aemter und der BA")
    zs["A5"] = "Zeitraum Rueckblick"
    zs["B5"] = f"ab {STARTYEAR}"
    for c in ("A2", "A3", "A4", "A5"):
        zs[c].font = Font(name=ARIAL, bold=True)
    zs["A7"] = ("Hinweis: Der Ausblick auf dem Blatt Perspektive ist keine "
                "Prognose. Trendfortschreibungen sind Illustrationen, "
                "belastbar ist vor allem die Demografie (Ersatzbedarf "
                "durch Verrentung).")
    zs["A8"] = ("Professionelle Referenzen zum Abgleich: QuBe-Projektionen "
                "von BIBB und IAB (bibb.de/qube) sowie das "
                "Fachkraeftemonitoring Baden-Wuerttemberg.")

    # ---------------- Arbeitslose (Monatsreihe ab STARTYEAR)
    if alo:
        m_alo = merkmal_wahl(alo, "arbeitslos")
        ab = wb.create_sheet("Arbeitslose")
        ab.append(["Monat"] + [KREISE[a] for a in ags_liste])
        werte = {(r["jahr"], r["monat"], r["ags"]): r["wert"]
                 for r in alo if r["merkmal"] == m_alo
                 and r["dim"] == "Insgesamt" and r["monat"]}
        monate = sorted({(j, m) for (j, m, _) in werte})
        for i, (j, m) in enumerate(monate, start=2):
            ab.cell(row=i, column=1,
                    value=date(j, m, 1)).number_format = "MM.YYYY"
            for k, a in enumerate(ags_liste):
                ab.cell(row=i, column=2 + k,
                        value=werte.get((j, m, a))).number_format = "#,##0"
        style_header(ab, 1 + len(ags_liste))
        nz = len(monate) + 1
        ch = flaeche3d(f"Arbeitslose je Kreis ab {STARTYEAR} (Monatswerte)")
        ch.x_axis.number_format = "YYYY"
        ch.add_data(Reference(ab, min_col=2, max_col=1 + len(ags_liste),
                              min_row=1, max_row=nz), titles_from_data=True)
        ch.set_categories(Reference(ab, min_col=1, min_row=2, max_row=nz))
        ab.add_chart(ch, f"A{nz + 3}")

    # ---------------- Beschaeftigte (Stichtage)
    if bes:
        m_bes = merkmal_wahl(bes, "besch")
        bb = wb.create_sheet("Beschaeftigte")
        bb.append(["Jahr"] + [KREISE[a] for a in ags_liste])
        werte = {}
        for r in bes:
            if r["merkmal"] != m_bes or r["dim"] != "Insgesamt":
                continue
            # bei mehreren Stichtagen je Jahr den letzten Wert nehmen
            werte[(r["jahr"], r["ags"])] = r["wert"]
        jahre = sorted({j for (j, _) in werte})
        for i, j in enumerate(jahre, start=2):
            bb.cell(row=i, column=1, value=j)
            for k, a in enumerate(ags_liste):
                bb.cell(row=i, column=2 + k,
                        value=werte.get((j, a))).number_format = "#,##0"
        style_header(bb, 1 + len(ags_liste))
        nz = len(jahre) + 1
        ch = flaeche3d("SV-Beschaeftigte je Kreis (Stichtage)")
        ch.add_data(Reference(bb, min_col=2, max_col=1 + len(ags_liste),
                              min_row=1, max_row=nz), titles_from_data=True)
        ch.set_categories(Reference(bb, min_col=1, min_row=2, max_row=nz))
        bb.add_chart(ch, f"A{nz + 3}")

    # ---------------- Demografie (Altersstruktur, letztes Jahr)
    bev_jahr = None
    if bev:
        m_bev = merkmal_wahl(bev, "bev")
        bev_jahre = sorted({r["jahr"] for r in bev if r["merkmal"] == m_bev})
        bev_jahr = bev_jahre[-1] if bev_jahre else None
    if bev_jahr:
        dg = wb.create_sheet("Demografie")
        dg.append([f"Altersgruppe ({bev_jahr})"]
                  + [KREISE[a] for a in ags_liste])
        gruppen = sorted(
            {r["dim"] for r in bev
             if r["jahr"] == bev_jahr and r["merkmal"] == m_bev
             and r["dim"] != "Insgesamt"},
            key=lambda g: (alter_grenzen(g) or (999, 999)))
        wmap = {(r["dim"], r["ags"]): r["wert"] for r in bev
                if r["jahr"] == bev_jahr and r["merkmal"] == m_bev}
        for i, g in enumerate(gruppen, start=2):
            dg.cell(row=i, column=1, value=g)
            for k, a in enumerate(ags_liste):
                dg.cell(row=i, column=2 + k,
                        value=wmap.get((g, a))).number_format = "#,##0"
        style_header(dg, 1 + len(ags_liste))

    # ---------------- Vorausberechnung
    if prog:
        m_prog = merkmal_wahl(prog, "bev")
        # falls Varianten enthalten sind, eine je Kreis einheitlich waehlen
        varianten = sorted({r["dim"] for r in prog
                            if r["merkmal"] == m_prog})
        variante = next((v for v in varianten if "haupt" in v.lower()),
                        varianten[0] if varianten else "Insgesamt")
        vp = wb.create_sheet("Vorausberechnung")
        vp.append([f"Jahr ({variante})"] + [KREISE[a] for a in ags_liste])
        werte = {(r["jahr"], r["ags"]): r["wert"] for r in prog
                 if r["merkmal"] == m_prog and r["dim"] == variante}
        jahre = sorted({j for (j, _) in werte})
        for i, j in enumerate(jahre, start=2):
            vp.cell(row=i, column=1, value=j)
            for k, a in enumerate(ags_liste):
                vp.cell(row=i, column=2 + k,
                        value=werte.get((j, a))).number_format = "#,##0"
        style_header(vp, 1 + len(ags_liste))
        nz = len(jahre) + 1
        ch = flaeche3d("Bevoelkerungsvorausberechnung je Kreis")
        ch.add_data(Reference(vp, min_col=2, max_col=1 + len(ags_liste),
                              min_row=1, max_row=nz), titles_from_data=True)
        ch.set_categories(Reference(vp, min_col=1, min_row=2, max_row=nz))
        vp.add_chart(ch, f"A{nz + 3}")

    # ---------------- Perspektive (Kennzahlen statt Kaffeesatz)
    pe = wb.create_sheet("Perspektive")
    pe.append(["Kennzahl"] + [KREISE[a] for a in ags_liste])
    style_header(pe, 1 + len(ags_liste))
    zeile = 2

    def schreib(name, werte_je_ags, fmt="#,##0"):
        nonlocal zeile
        pe.cell(row=zeile, column=1, value=name)
        for k, a in enumerate(ags_liste):
            z = pe.cell(row=zeile, column=2 + k, value=werte_je_ags.get(a))
            z.number_format = fmt
        zeile += 1

    if alo:
        m_alo = merkmal_wahl(alo, "arbeitslos")
        reihen = {}
        for r in alo:
            if (r["merkmal"] == m_alo and r["dim"] == "Insgesamt"
                    and r["monat"]):
                reihen.setdefault(r["ags"], []).append(
                    ((r["jahr"], r["monat"]), r["wert"]))
        d12, dvor = {}, {}
        for a, punkte in reihen.items():
            punkte.sort()
            letzte = [w for _, w in punkte[-12:] if w is not None]
            basis = [w for (j, _), w in punkte
                     if 2010 <= j <= 2019 and w is not None]
            if letzte:
                d12[a] = round(sum(letzte) / len(letzte))
            if letzte and basis:
                dvor[a] = (sum(letzte) / len(letzte)) \
                    / (sum(basis) / len(basis)) - 1
        schreib("Arbeitslose, Mittel letzte 12 Monate", d12)
        schreib("dito vs. Mittel 2010-2019", dvor, "+0.0%;-0.0%")

    if bes:
        m_bes = merkmal_wahl(bes, "besch")
        bj = {}
        for r in bes:
            if r["merkmal"] == m_bes and r["dim"] == "Insgesamt":
                bj[(r["ags"], r["jahr"])] = r["wert"]
        jahre = sorted({j for (_, j) in bj})
        if len(jahre) >= 2:
            j_neu = jahre[-1]
            j_alt = max(j for j in jahre if j <= j_neu - 10) \
                if any(j <= j_neu - 10 for j in jahre) else jahre[0]
            cagr, p5, p10 = {}, {}, {}
            for a in ags_liste:
                w0, w1 = bj.get((a, j_alt)), bj.get((a, j_neu))
                if w0 and w1 and j_neu > j_alt:
                    g = (w1 / w0) ** (1 / (j_neu - j_alt)) - 1
                    cagr[a] = g
                    p5[a] = round(w1 * (1 + g) ** 5)
                    p10[a] = round(w1 * (1 + g) ** 10)
            schreib(f"Beschaeftigte, Wachstum p.a. {j_alt}-{j_neu}",
                    cagr, "+0.0%;-0.0%")
            schreib(f"Trendfortschreibung {j_neu + 5} (Illustration)", p5)
            schreib(f"Trendfortschreibung {j_neu + 10} (Illustration)", p10)

    if bev_jahr:
        m_bev = merkmal_wahl(bev, "bev")
        nach, abg, senior = {}, {}, {}
        quot = {}
        for a in ags_liste:
            n = summe_altersband(bev, bev_jahr, a, m_bev, 15, 25)
            g = summe_altersband(bev, bev_jahr, a, m_bev, 55, 65)
            s = summe_altersband(bev, bev_jahr, a, m_bev, 65, 200)
            if n is not None:
                nach[a] = n
            if g is not None:
                abg[a] = g
            if s is not None:
                senior[a] = s
            if n and g:
                quot[a] = n / g
        schreib(f"Bevoelkerung 15-24 ({bev_jahr})", nach)
        schreib(f"Bevoelkerung 55-64 ({bev_jahr})", abg)
        schreib(f"Bevoelkerung 65+ ({bev_jahr})", senior)
        schreib("Nachwuchsquotient (15-24 je 55-64)", quot, "0.00")

    if prog:
        m_prog = merkmal_wahl(prog, "bev")
        varianten = sorted({r["dim"] for r in prog
                            if r["merkmal"] == m_prog})
        variante = next((v for v in varianten if "haupt" in v.lower()),
                        varianten[0] if varianten else "Insgesamt")
        pw = {(r["ags"], r["jahr"]): r["wert"] for r in prog
              if r["merkmal"] == m_prog and r["dim"] == variante}
        jahre = sorted({j for (_, j) in pw})
        basisj = max((j for j in jahre if j <= heute.year),
                     default=jahre[0] if jahre else None)
        for ziel in (heute.year + 5, heute.year + 10):
            zj = min((j for j in jahre if j >= ziel), default=None)
            if not (basisj and zj and zj > basisj):
                continue
            d = {}
            for a in ags_liste:
                w0, w1 = pw.get((a, basisj)), pw.get((a, zj))
                if w0 and w1:
                    d[a] = w1 / w0 - 1
            schreib(f"Bevoelkerung {basisj} -> {zj} (Vorausber.)",
                    d, "+0.0%;-0.0%")

    zeile += 1
    pe.cell(row=zeile, column=1, value=(
        "Einordnung: Belastbar ist der demografische Ersatzbedarf "
        "(Jahrgaenge 55-64 gehen binnen 10 Jahren in Rente, der "
        "Nachwuchsquotient zeigt, wie viel nachrueckt). "
        "Trendfortschreibungen ignorieren Konjunktur und Schocks. "
        "Berufsscharfe Projektionen: QuBe (BIBB/IAB) und "
        "Fachkraeftemonitoring BW heranziehen."))

    fehlend = [k for k in TABELLEN if not (DATA / f"{k}.csv").exists()]
    if fehlend:
        zeile += 2
        pe.cell(row=zeile, column=1, value=(
            "Noch keine Daten fuer: " + ", ".join(fehlend)
            + " (erster Lauf ausstehend oder Tabellencode pruefen, "
              "siehe Actions-Log)."))

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font is None or cell.font.name is None:
                    cell.font = Font(name=ARIAL)
        for col in ws.columns:
            breite = max((len(str(c.value)) for c in col
                          if c.value is not None), default=0)
            ws.column_dimensions[
                get_column_letter(col[0].column)].width = min(
                max(breite + 2, 10), 60)
    wb.save(OUT)
    print(f"Excel geschrieben: {OUT}")
    return 0


if __name__ == "__main__":
    fetch_all()
    raise SystemExit(build_excel())
