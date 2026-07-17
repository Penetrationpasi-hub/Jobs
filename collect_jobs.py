#!/usr/bin/env python3
"""
BA-Jobmonitor: Sammelt taeglich Stellenangebote der Bundesagentur fuer Arbeit
fuer Waiblingen, Rems-Murr-Kreis und Ostalbkreis ueber die Jobsuche-API.

Auth: X-API-Key: jobboerse-jobsuche (oeffentliche clientId, keine Registrierung noetig)
API-Doku: https://github.com/bundesAPI/jobsuche-api

Outputs (im Ordner data/):
  snapshots.csv   - ein Datensatz pro Tag und Region (Gesamtzahlen, Verteilungen)
  berufe.csv      - Anzahl Stellen pro Beruf, Tag und Region (Top-Auswertung)
  jobs_log.csv    - jede jemals gesehene Stelle mit first_seen/last_seen
                    -> ermoeglicht Laufzeit- und Fluktuationsanalysen
  arbeitslose.csv - monatliche Arbeitslose + Quote je Kreis aus der
                    Regionaldatenbank Deutschland (GENESIS, EVAS 13211).
                    Benoetigt env GENESIS_USER / GENESIS_PASS
                    (kostenloses Konto auf regionalstatistik.de).
"""

import csv
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date
from pathlib import Path

API_BASE = "https://rest.arbeitsagentur.de/jobboerse/jobsuche-service/pc/v4/jobs"
HEADERS = {
    "X-API-Key": "jobboerse-jobsuche",
    "User-Agent": "ba-jobmonitor/1.0 (privates Arbeitsmarkt-Monitoring)",
}

# Regionen: Suche ueber Ort + Umkreis in km.
# Die API kennt keinen sauberen Landkreis-Filter, daher Naeherung ueber Umkreise.
# Waiblingen eng (15 km) als Kernregion, die Kreise grob ueber zentrale Orte.
REGIONS = {
    "waiblingen_30km": {"wo": "Waiblingen", "umkreis": 30},
    "rems-murr-kreis": {"wo": "Backnang", "umkreis": 25},
    "ostalbkreis": {"wo": "Aalen", "umkreis": 30},
    "backnang_10km": {"wo": "Backnang", "umkreis": 10},
    # Ostseite Stuttgarts: die API kennt nur Ort + Umkreis, keine
    # Stadtbezirke. Naeherung ueber Fellbach (grenzt direkt an
    # Bad Cannstatt/Untertuerkheim), kleiner Radius.
    "stuttgart_ost": {"wo": "Fellbach", "umkreis": 8},
}

PAGE_SIZE = 100
# Sicherheitslimit je Suchgebiet: 300 Seiten x 100 = 30000 Stellen.
# Ein hartes API-Limit ist nicht dokumentiert; macht die API frueher
# dicht (Fehler oder dauerhaft leere Seiten), bricht fetch_region
# sauber ab und behaelt das Teilergebnis.
MAX_PAGES = 300
REQUEST_PAUSE = 1.0     # Sekunden zwischen Requests, hoeflich bleiben
DATA_DIR = Path(__file__).parent / "data"

# ---------------- Regionaldatenbank Deutschland (GENESIS, EVAS 13211)
# Monatliche Arbeitslose und Arbeitslosenquoten auf Kreisebene.
# Kostenloses Konto auf regionalstatistik.de noetig, Zugangsdaten als
# GitHub Secrets GENESIS_USER / GENESIS_PASS. Nur POST, GET ist abgeschaltet.
GENESIS_BASE = "https://www.regionalstatistik.de/genesisWS/rest/2020"
# Kreistabelle Arbeitslose/Quoten, Monatswerte. Falls der Code nicht passt,
# listet fetch_arbeitslose() den Katalog und man setzt env AL_TABLE.
AL_TABLE = os.environ.get("AL_TABLE", "13211-01-03-4-B")
AL_KREISE = {
    "08119": "Rems-Murr-Kreis",
    "08136": "Ostalbkreis",
    "08111": "Stuttgart",         # fuer das Suchgebiet Stuttgart Ost
}
# Suchgebiet -> AGS fuer die Kreis-Relation im Excel
REGION_ZU_AGS = {
    "rems-murr-kreis": "08119",
    "ostalbkreis": "08136",
    "waiblingen_30km": "08119",   # Waiblingen liegt im Rems-Murr-Kreis
    "backnang_10km": "08119",
    "stuttgart_ost": "08111",     # Naeherung, Suchgebiet ist Fellbach 8 km
}
AL_CSV = DATA_DIR / "arbeitslose.csv"


def genesis_creds() -> tuple[str, str] | None:
    user = os.environ.get("GENESIS_USER", "").strip()
    pw = os.environ.get("GENESIS_PASS", "").strip()
    if not user or not pw:
        return None
    return user, pw


def genesis_post(endpoint: str, params: dict) -> bytes:
    """POST an die GENESIS-REST-Schnittstelle, Zugangsdaten im Header."""
    user, pw = genesis_creds()
    body = urllib.parse.urlencode(params).encode("utf-8")
    req = urllib.request.Request(
        f"{GENESIS_BASE}/{endpoint}", data=body, method="POST",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "username": user,
            "password": pw,
            "User-Agent": HEADERS["User-Agent"],
        })
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


MONAT_NUM = {f"MONAT{m:02d}": m for m in range(1, 13)}


def parse_zahl(s):
    s = str(s or "").strip()
    if s in ("", "-", ".", "...", "x", "/"):
        return None
    try:
        f = float(s.replace(",", "."))
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


def fetch_arbeitslose():
    """Monatliche Arbeitslose je Kreis aus GENESIS ziehen -> arbeitslose.csv.

    Voller Neuabruf ab Vorjahr bei jedem Lauf (Daten werden revidiert).
    Ohne Zugangsdaten wird der Schritt uebersprungen, das Excel nutzt
    dann den letzten Stand der CSV.
    """
    if genesis_creds() is None:
        print("GENESIS_USER/GENESIS_PASS nicht gesetzt, "
              "Arbeitslosen-Abruf uebersprungen.", file=sys.stderr)
        return
    jahr = date.today().year
    params = {
        "name": AL_TABLE,
        "area": "free",
        "regionalvariable": "KREISE",
        "regionalkey": ",".join(AL_KREISE),
        "startyear": jahr - 1,
        "endyear": jahr,
        "format": "ffcsv",
        "language": "de",
        "compress": "false",
    }
    try:
        raw = genesis_post("data/tablefile", params)
    except Exception as e:
        print(f"GENESIS-Abruf fehlgeschlagen: {e}", file=sys.stderr)
        _genesis_katalog_hinweis()
        return
    text = raw.decode("utf-8-sig", errors="replace")
    if text.lstrip().startswith("{"):
        # Fehlerantwort kommt als JSON (z.B. falscher Tabellencode/Login)
        print(f"GENESIS-Fehlerantwort: {text[:400]}", file=sys.stderr)
        _genesis_katalog_hinweis()
        return

    zeilen = {}   # (jahr, monat, ags) -> {arbeitslose, quote}
    reader = csv.DictReader(text.splitlines(), delimiter=";")
    feld = reader.fieldnames or []
    # Wertspalten tolerant erkennen (Benennungen variieren je Tabelle)
    f_alo = next((f for f in feld
                  if "Arbeitslose" in f and "Anzahl" in f), None)
    f_quote = next((f for f in feld if "quote" in f.lower()
                    and "zivilen" in f.lower()), None) \
        or next((f for f in feld if "quote" in f.lower()), None)
    if not f_alo:
        print(f"Wertspalte nicht gefunden. Spalten: {feld}", file=sys.stderr)
        return
    for row in reader:
        ags = (row.get("1_Auspraegung_Code") or "").strip()
        if ags not in AL_KREISE:
            continue
        try:
            j = int((row.get("Zeit") or "0").strip())
        except ValueError:
            continue
        m = MONAT_NUM.get((row.get("2_Auspraegung_Code") or "").strip())
        if not j or not m:
            continue
        alo = parse_zahl(row.get(f_alo, ""))
        if alo is None:
            continue
        quote = parse_zahl(row.get(f_quote, "")) if f_quote else None
        zeilen[(j, m, ags)] = {"arbeitslose": alo, "quote": quote}

    if not zeilen:
        print("GENESIS: keine verwertbaren Zeilen erhalten.", file=sys.stderr)
        return
    # Bestehende CSV einlesen und mit neuen Monaten zusammenfuehren,
    # damit Historie vor dem Abruffenster erhalten bleibt
    if AL_CSV.exists():
        with AL_CSV.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                k = (int(row["jahr"]), int(row["monat"]), row["ags"])
                zeilen.setdefault(k, {
                    "arbeitslose": parse_zahl(row["arbeitslose"]),
                    "quote": parse_zahl(row["quote"]),
                })
    with AL_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["jahr", "monat", "ags", "kreis", "arbeitslose", "quote"])
        for (j, m, ags) in sorted(zeilen):
            v = zeilen[(j, m, ags)]
            w.writerow([j, m, ags, AL_KREISE[ags],
                        v["arbeitslose"],
                        "" if v["quote"] is None else v["quote"]])
    print(f"Arbeitslosenzahlen aktualisiert: {len(zeilen)} Zeilen "
          f"({AL_CSV.name}).")


def _genesis_katalog_hinweis():
    """Bei Problemen verfuegbare 13211er-Tabellen listen (Diagnose)."""
    try:
        raw = genesis_post("catalogue/tables", {
            "selection": "13211*", "area": "free",
            "pagelength": 60, "language": "de",
        })
        data = json.loads(raw.decode("utf-8", errors="replace"))
        for t in (data.get("List") or [])[:60]:
            print(f"  Tabelle: {t.get('Code')} - {t.get('Content')}",
                  file=sys.stderr)
        print("Falls noetig: passenden Code als env AL_TABLE setzen.",
              file=sys.stderr)
    except Exception as e:
        print(f"Katalogabruf fehlgeschlagen: {e}", file=sys.stderr)


def lade_arbeitslose():
    """arbeitslose.csv einlesen -> Liste sortierter Monatszeilen."""
    if not AL_CSV.exists():
        return []
    rows = []
    with AL_CSV.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append({
                "jahr": int(row["jahr"]), "monat": int(row["monat"]),
                "ags": row["ags"], "kreis": row["kreis"],
                "arbeitslose": parse_zahl(row["arbeitslose"]),
                "quote": parse_zahl(row["quote"]),
            })
    rows.sort(key=lambda r: (r["jahr"], r["monat"], r["ags"]))
    return rows


def fetch_page(params: dict) -> dict | None:
    """Eine Ergebnisseite laden. Liefert None statt einer Exception,
    wenn nach 3 Versuchen nichts kommt -- der Aufrufer entscheidet, ob
    das ein API-Limit (Teilergebnis behalten) oder ein Ausfall ist."""
    qs = urllib.parse.urlencode(params)
    req = urllib.request.Request(f"{API_BASE}?{qs}", headers=HEADERS)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            print(f"  Fehler (Versuch {attempt + 1}/3): {e}", file=sys.stderr)
            time.sleep(5 * (attempt + 1))
    print(f"  Aufgegeben: {params.get('wo')} Seite {params.get('page')}",
          file=sys.stderr)
    return None


def fetch_region(region_params: dict) -> tuple[list[dict], int]:
    """Alle Stellen einer Region einsammeln (paginiert).

    Tolerant gegen ein API-seitiges Paginierungslimit: schlaegt eine
    Seite > 1 dauerhaft fehl, wird abgebrochen und das Teilergebnis
    behalten. Nur wenn schon Seite 1 scheitert, fliegt eine Exception.
    Dubletten ueber Seitengrenzen (die Sortierung kann bei laufenden
    Aenderungen verrutschen) werden per refnr uebersprungen.
    """
    jobs = []
    refs: set[str] = set()
    max_ergebnisse = 0
    for page in range(1, MAX_PAGES + 1):
        params = {
            "angebotsart": 1,        # 1 = Arbeit
            "page": page,
            "size": PAGE_SIZE,
            "pav": "false",          # keine privaten Arbeitsvermittler
            **region_params,
        }
        data = fetch_page(params)
        if data is None:
            if page == 1:
                raise RuntimeError(f"Seite 1 nicht abrufbar: {region_params}")
            print(f"  Stopp bei Seite {page} (moegliches API-Limit), "
                  f"behalte {len(jobs)} Stellen.", file=sys.stderr)
            break
        max_ergebnisse = int(data.get("maxErgebnisse", 0)) or max_ergebnisse
        batch = data.get("stellenangebote", [])
        if not batch:
            break
        for j in batch:
            ref = j.get("refnr") or j.get("referenznummer")
            if ref:
                if ref in refs:
                    continue
                refs.add(ref)
            jobs.append(j)
        if max_ergebnisse and len(jobs) >= max_ergebnisse:
            break
        time.sleep(REQUEST_PAUSE)
    return jobs, max_ergebnisse


def append_csv(path: Path, fieldnames: list[str], rows: list[dict]):
    exists = path.exists()
    with path.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if not exists:
            w.writeheader()
        w.writerows(rows)


def update_jobs_log(path: Path, today: str, seen: dict[str, dict]):
    """Stellen-Log fortschreiben: neue Stellen anlegen, bekannte mit last_seen aktualisieren."""
    fieldnames = ["refnr", "region", "beruf", "titel", "arbeitgeber",
                  "ort", "first_seen", "last_seen"]
    existing: dict[str, dict] = {}
    if path.exists():
        with path.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing[row["refnr"]] = row

    for refnr, info in seen.items():
        if refnr in existing:
            existing[refnr]["last_seen"] = today
        else:
            existing[refnr] = {**info, "first_seen": today, "last_seen": today}

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(existing.values())

    return existing


def main():
    DATA_DIR.mkdir(exist_ok=True)
    today = date.today().isoformat()
    snapshot_rows = []
    berufe_rows = []
    seen_jobs: dict[str, dict] = {}

    for region, rparams in REGIONS.items():
        print(f"[{region}] Abruf laeuft ...")
        try:
            jobs, total = fetch_region(rparams)
        except Exception as e:
            # Region ueberspringen statt den ganzen Lauf zu reissen; der
            # Tag bleibt in der Zeitreihe dann leer (kein falsches 0)
            print(f"[{region}] Abruf fehlgeschlagen, uebersprungen: {e}",
                  file=sys.stderr)
            continue
        print(f"[{region}] {len(jobs)} von {total} Stellen geladen")
        if total > len(jobs):
            print(f"[{region}] Hinweis: Detaildaten unvollstaendig "
                  f"({len(jobs)}/{total}).", file=sys.stderr)

        berufe = Counter()
        arbeitszeit = Counter()
        befristung = Counter()
        arbeitgeber = Counter()

        for j in jobs:
            beruf = (j.get("beruf") or "unbekannt").strip()
            berufe[beruf] += 1
            arbeitgeber[(j.get("arbeitgeber") or "unbekannt").strip()] += 1
            ort = ""
            if isinstance(j.get("arbeitsort"), dict):
                ort = j["arbeitsort"].get("ort") or ""
            refnr = j.get("refnr") or j.get("referenznummer")
            if refnr and refnr not in seen_jobs:
                seen_jobs[refnr] = {
                    "refnr": refnr,
                    "region": region,
                    "beruf": beruf,
                    "titel": (j.get("titel") or "").strip()[:200],
                    "arbeitgeber": (j.get("arbeitgeber") or "").strip()[:150],
                    "ort": ort,
                }

        snapshot_rows.append({
            "datum": today,
            "region": region,
            "stellen_gesamt": total,
            "stellen_geladen": len(jobs),
            "berufe_unique": len(berufe),
            "arbeitgeber_unique": len(arbeitgeber),
            "top_beruf": berufe.most_common(1)[0][0] if berufe else "",
            "top_arbeitgeber": arbeitgeber.most_common(1)[0][0] if arbeitgeber else "",
        })

        for beruf, n in berufe.most_common():
            berufe_rows.append({
                "datum": today, "region": region, "beruf": beruf, "anzahl": n,
            })

    append_csv(DATA_DIR / "snapshots.csv",
               ["datum", "region", "stellen_gesamt", "stellen_geladen",
                "berufe_unique", "arbeitgeber_unique", "top_beruf", "top_arbeitgeber"],
               snapshot_rows)
    append_csv(DATA_DIR / "berufe.csv",
               ["datum", "region", "beruf", "anzahl"],
               berufe_rows)
    log = update_jobs_log(DATA_DIR / "jobs_log.csv", today, seen_jobs)

    fetch_arbeitslose()

    print(f"\nFertig. Snapshots: {len(snapshot_rows)} Regionen, "
          f"Stellen-Log: {len(log)} Stellen insgesamt bekannt.")




# ==================================================================
# Teil 2: Auswertung -> arbeitsmarkt_auswertung.xlsx
# ==================================================================

import csv
import os
import statistics
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")
OUT = os.path.join(BASE, "arbeitsmarkt_auswertung.xlsx")

REGION_NAMEN = {
    "waiblingen_15km": "Waiblingen 15 km (alt)",
    "waiblingen_30km": "Waiblingen 30 km",
    "rems-murr-kreis": "Rems-Murr-Kreis",
    "ostalbkreis": "Ostalbkreis",
    "backnang_10km": "Backnang 10 km",
    "stuttgart_ost": "Stuttgart Ost (Fellbach 8 km)",
}

# ---- Umfang der Blaetter. None = unbegrenzt (alle Eintraege). ------
# Aktuell, Standzeit, Arbeitgeber und Arbeitgeber-Tempo zeigen ALLES.
# Berufe-Zeitreihe legt je Beruf eine eigene SPALTE an und bleibt
# deshalb gedeckelt (None geht auch, macht das Blatt nur sehr breit).
# AL-Relation ist das Blatt mit manueller Eingabe, dort reichen 20.
TOP_BERUFE_ZEITREIHE = 50
TOP_AL_RELATION = 20
# Arbeitgeber-Tempo: mindestens so viele beendete Stellen noetig,
# sonst dominieren Einzelfaelle die Rangliste
AG_TEMPO_MIN_BEENDET = 5

ARIAL = "Arial"
HEAD_FONT = Font(name=ARIAL, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", start_color="1D3C6E")
INPUT_FONT = Font(name=ARIAL, color="0000FF")
DATUM_FMT = "DD.MM.YYYY"


def read_csv(name):
    path = os.path.join(DATA, name)
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def as_date(s):
    return datetime.strptime(s.strip(), "%Y-%m-%d").date()


def standzeit_tage(r):
    """Standzeit einer Log-Zeile in Tagen (erster Tag zaehlt mit)."""
    return (as_date(r["last_seen"]) - as_date(r["first_seen"])).days + 1


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    if row == 1:
        ws.freeze_panes = "A2"


def autosize(ws, maxw=45):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, maxw)


def base_font(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.font is None or c.font.name != ARIAL:
                c.font = Font(name=ARIAL, bold=c.font.bold if c.font else False,
                              color=c.font.color if c.font else None)


# ---- Diagramm-Helfer: 2D statt der alten 3D-Flaechen. --------------
# Bei den AreaChart3D verdeckte die vorderste Flaeche alle kleineren
# Reihen, und die Zeitachse war in der 3D-Perspektive unlesbar.

def _achsen_sichtbar(ch, x_titel=None, y_titel=None,
                     x_fmt=None, y_fmt="#,##0"):
    # Excel unterdrueckt Achsen, wenn delete nicht explizit False ist
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    if x_titel:
        ch.x_axis.title = x_titel
    if y_titel:
        ch.y_axis.title = y_titel
    if x_fmt:
        ch.x_axis.number_format = x_fmt
    if y_fmt:
        ch.y_axis.number_format = y_fmt
    ch.y_axis.majorGridlines = ChartLines()
    ch.legend.position = "b"


def linien_chart(titel, x_titel, y_titel, x_fmt, n_kategorien,
                 hoehe=10, breite=22):
    """Liniendiagramm fuer Zeitreihen: jede Reihe bleibt sichtbar."""
    ch = LineChart()
    ch.title = titel
    ch.height, ch.width = hoehe, breite
    _achsen_sichtbar(ch, x_titel, y_titel, x_fmt)
    # Datumsbeschriftung ausduennen, sonst wird die Achse bei vielen
    # Erhebungstagen unleserlich (~12 Beschriftungen sind gut lesbar)
    skip = max(1, n_kategorien // 12)
    ch.x_axis.tickLblSkip = skip
    ch.x_axis.tickMarkSkip = skip
    return ch


def balken_chart(titel, wert_titel, n_zeilen, hoehe=None, breite=22):
    """Horizontales Balkendiagramm: lange Namen bleiben lesbar,
    kleine Werte verschwinden nicht hinter grossen."""
    ch = BarChart()
    ch.type = "bar"
    ch.grouping = "clustered"
    ch.gapWidth = 60
    ch.title = titel
    ch.height = hoehe if hoehe else max(8.0, min(26.0, 1 + 0.75 * n_zeilen))
    ch.width = breite
    _achsen_sichtbar(ch, None, wert_titel)
    # Rang 1 oben statt unten (Excel zeichnet Kategorien von unten nach
    # oben); Wertachse dabei unten am Diagramm lassen
    ch.x_axis.scaling.orientation = "maxMin"
    ch.y_axis.crosses = "max"
    return ch


def build_excel():
    berufe = read_csv("berufe.csv")
    snaps = read_csv("snapshots.csv")
    log = read_csv("jobs_log.csv")
    if not berufe or not snaps:
        print("FEHLER: keine Daten gefunden, zuerst collect_jobs.py laufen lassen.")
        return 1

    daten = sorted({r["datum"] for r in berufe})
    stand = daten[-1]
    stand_d = as_date(stand)
    snap_map = {(s["datum"], s["region"]): int(s["stellen_gesamt"]) for s in snaps}
    # Aktive Suchgebiete (am Stichtag erhoben) zuerst, eingestellte ans Ende
    alle_reg = sorted({r["region"] for r in berufe} | {s["region"] for s in snaps})
    aktiv = [r for r in alle_reg if (stand, r) in snap_map]
    regionen = aktiv + [r for r in alle_reg if r not in aktiv]
    reg_name = lambda r: REGION_NAMEN.get(r, r)

    # Vergleichsdaten: naechster vorhandener Stichtag ca. 7 bzw. 30 Tage zurueck
    def vergleich(tage):
        ziel = [d for d in daten if (stand_d - as_date(d)).days >= tage]
        return ziel[-1] if ziel else None
    d7, d30 = vergleich(7), vergleich(30)

    # Bestaende je (datum, region, beruf) und je (datum, beruf) gesamt
    brb = {}
    for r in berufe:
        brb[(r["datum"], r["region"], r["beruf"])] = int(r["anzahl"])
    # ueber brb (dedupliziert) aggregieren: schuetzt vor Doppelzaehlung,
    # falls die Erhebung an einem Tag mehrfach lief
    bges = {}
    for (d, _rg, b), n in brb.items():
        bges[(d, b)] = bges.get((d, b), 0) + n

    top_akt = sorted((b for (d, b) in bges if d == stand),
                     key=lambda b: -bges[(stand, b)])
    # frueher Top 20/40 -- jetzt zeigt Aktuell ALLE Berufe des Stichtags,
    # nur das Spalten-Blatt und das Eingabe-Blatt bleiben begrenzt
    bz_berufe = top_akt if TOP_BERUFE_ZEITREIHE is None \
        else top_akt[:TOP_BERUFE_ZEITREIHE]
    al_berufe = top_akt if TOP_AL_RELATION is None \
        else top_akt[:TOP_AL_RELATION]

    wb = Workbook()

    # ---------------- Zusammenfassung
    zs = wb.active
    zs.title = "Zusammenfassung"
    zs["A1"] = "Arbeitsmarkt-Auswertung Rems-Murr / Ostalb"
    zs["A1"].font = Font(name=ARIAL, bold=True, size=14)
    zs["A2"] = "Stand"
    zs["B2"] = stand_d
    zs["B2"].number_format = DATUM_FMT
    zs["A3"] = "Erhebungstage"
    zs["B3"] = len(daten)
    zs["A4"] = "Regionen"
    zs["B4"] = ", ".join(reg_name(r) for r in regionen)
    zs["A5"] = "Stellen je gesehen (Log)"
    zs["B5"] = len(log)
    zs["A7"] = (f"Hinweis: Das Script laedt bis zu {MAX_PAGES * PAGE_SIZE} "
                "Stellen je Suchgebiet im Detail (bei API-Abbruch frueher, "
                "siehe Actions-Log). Liegt 'Stellen gesamt' darueber, sind "
                "Berufs- und Arbeitgeberzahlen eine Stichprobe; die Spalte "
                "'Stellen gesamt' in der Zeitreihe ist exakt.")
    zs["A8"] = ("Arbeitslose je Beruf fuer das Blatt AL-Relation weiterhin "
                "manuell aus statistik.arbeitsagentur.de eintragen (blaue "
                "Spalte). Kreiszahlen kommen automatisch aus der "
                "Regionaldatenbank Deutschland (Blatt Arbeitslose).")
    zs["A9"] = ("Hinweis: Die Suchgebiete ueberschneiden sich (Umkreissuchen), "
                "'Summe Suchgebiete' zaehlt Stellen daher mehrfach. Die Summe "
                "laeuft nur ueber aktive Gebiete; eingestellte Suchgebiete "
                "stehen in der Zeitreihe rechts neben der Summenspalte.")
    zs["A10"] = ("Alle Zahlen sind beim Erzeugen fest berechnet (keine "
                 "Formeln), damit sie auch in Vorschau-Apps und auf dem "
                 "Handy angezeigt werden. Einzige Ausnahme: die Relation "
                 "zur blauen Eingabespalte in AL-Relation rechnet live.")
    for c in ("A2", "A3", "A4", "A5"):
        zs[c].font = Font(name=ARIAL, bold=True)

    # ---------------- Zeitreihe (snapshots: stellen_gesamt exakt)
    zt = wb.create_sheet("Zeitreihe")
    n_akt = len(aktiv)
    alt_reg = regionen[n_akt:]
    zt.append(["Datum"] + [reg_name(r) for r in aktiv] + ["Summe Suchgebiete"]
              + [reg_name(r) for r in alt_reg])
    for i, d in enumerate(daten, start=2):
        zt.cell(row=i, column=1, value=as_date(d)).number_format = DATUM_FMT
        werte = []
        for j, r in enumerate(aktiv, start=2):
            v = snap_map.get((d, r))
            zt.cell(row=i, column=j, value=v)
            if v is not None:
                werte.append(v)
        # Summe nur ueber aktive Gebiete -> Reihe bleibt ueber die Zeit
        # vergleichbar, auch wenn ein Suchgebiet eingestellt wurde
        zt.cell(row=i, column=n_akt + 2, value=sum(werte) if werte else None)
        for j, r in enumerate(alt_reg, start=n_akt + 3):
            zt.cell(row=i, column=j, value=snap_map.get((d, r)))
    style_header(zt, len(regionen) + 2)
    for row in zt.iter_rows(min_row=2, min_col=2):
        for c in row:
            c.number_format = "#,##0"
    ch = linien_chart("Offene Stellen je Suchgebiet", "Datum", "Stellen",
                      "DD.MM.", len(daten))
    # nur die Suchgebiete als Linien; die Summe wuerde die Skala strecken
    # und die einzelnen Gebiete optisch plattdruecken
    ch.add_data(Reference(zt, min_col=2, max_col=n_akt + 1,
                          min_row=1, max_row=len(daten) + 1),
                titles_from_data=True)
    ch.set_categories(Reference(zt, min_col=1, min_row=2,
                                max_row=len(daten) + 1))
    zt.add_chart(ch, f"A{len(daten) + 4}")

    # ---------------- Berufe-Zeitreihe (je Beruf eine Spalte, deshalb
    #                  auf TOP_BERUFE_ZEITREIHE begrenzt)
    bz = wb.create_sheet("Berufe-Zeitreihe")
    bz.append(["Datum"] + bz_berufe)
    for i, d in enumerate(daten, start=2):
        bz.cell(row=i, column=1, value=as_date(d)).number_format = DATUM_FMT
        for j, b in enumerate(bz_berufe, start=2):
            bz.cell(row=i, column=j, value=bges.get((d, b)))
    style_header(bz, len(bz_berufe) + 1)

    # ---------------- Aktuell (alle Berufe des Stichtags)
    ak = wb.create_sheet("Aktuell")
    kopf = (["Beruf"] + [reg_name(r) for r in regionen] +
            ["Gesamt", f"Vor 7 Tagen ({d7 or 'n/a'})",
             f"Vor 30 Tagen ({d30 or 'n/a'})", "Δ 7 Tage", "Δ 30 Tage"])
    ak.append(kopf)
    nr = len(regionen)
    for i, b in enumerate(top_akt, start=2):
        ak.cell(row=i, column=1, value=b)
        for j, r in enumerate(regionen, start=2):
            ak.cell(row=i, column=j, value=brb.get((stand, r, b)))
        ges = bges.get((stand, b), 0)
        v7 = bges.get((d7, b)) if d7 else None
        v30 = bges.get((d30, b)) if d30 else None
        ak.cell(row=i, column=nr + 2, value=ges)
        ak.cell(row=i, column=nr + 3, value=v7)
        ak.cell(row=i, column=nr + 4, value=v30)
        ak.cell(row=i, column=nr + 5,
                value=(ges - v7) if v7 is not None else None)
        ak.cell(row=i, column=nr + 6,
                value=(ges - v30) if v30 is not None else None)
    style_header(ak, nr + 6)
    n_chart = min(15, len(top_akt))
    fl = balken_chart(f"Top {n_chart} Berufe nach Suchgebiet (aktuell)",
                      "Stellen", n_chart, hoehe=18, breite=24)
    # nur aktive Suchgebiete als Serien (eingestellte Spalten sind leer)
    fl.add_data(Reference(ak, min_col=2, max_col=n_akt + 1,
                          min_row=1, max_row=n_chart + 1),
                titles_from_data=True)
    fl.set_categories(Reference(ak, min_col=1, min_row=2,
                                max_row=n_chart + 1))
    ak.add_chart(fl, f"A{len(top_akt) + 4}")

    # ---------------- Log (Rohdaten)
    lg = wb.create_sheet("Log")
    lg.append(["Refnr", "Region", "Beruf", "Arbeitgeber",
               "Erstmals gesehen", "Zuletzt gesehen", "Standzeit (Tage)"])
    for i, r in enumerate(log, start=2):
        lg.cell(row=i, column=1, value=r["refnr"])
        lg.cell(row=i, column=2, value=reg_name(r["region"]))
        lg.cell(row=i, column=3, value=r["beruf"])
        lg.cell(row=i, column=4, value=r["arbeitgeber"])
        lg.cell(row=i, column=5, value=as_date(r["first_seen"])).number_format = DATUM_FMT
        lg.cell(row=i, column=6, value=as_date(r["last_seen"])).number_format = DATUM_FMT
        lg.cell(row=i, column=7, value=standzeit_tage(r))
    style_header(lg, 7)

    # ---------------- Standzeit je Beruf (ALLE Berufe im Log,
    #                  sortiert nach Zahl der Log-Eintraege)
    # Fest berechnet statt COUNTIFS: bei unbegrenzter Berufsliste und
    # grossem Log wuerde Excel sonst bei jedem Oeffnen lange rechnen,
    # und in Vorschau-Apps blieben Formelzellen leer.
    beruf_stat = {}
    for r in log:
        s = beruf_stat.setdefault(r["beruf"],
                                  {"log": 0, "offen": 0, "dauern": []})
        s["log"] += 1
        if r["last_seen"] == stand:
            s["offen"] += 1
        else:
            s["dauern"].append(standzeit_tage(r))
    st = wb.create_sheet("Standzeit")
    st.append(["Beruf", "Stellen im Log", "Aktuell offen", "Beendet",
               "Ø Standzeit beendeter Stellen (Tage)"])
    st_berufe = sorted(beruf_stat, key=lambda b: -beruf_stat[b]["log"])
    for i, b in enumerate(st_berufe, start=2):
        s = beruf_stat[b]
        st.cell(row=i, column=1, value=b)
        st.cell(row=i, column=2, value=s["log"])
        st.cell(row=i, column=3, value=s["offen"])
        st.cell(row=i, column=4, value=len(s["dauern"]))
        if s["dauern"]:
            st.cell(row=i, column=5,
                    value=round(statistics.mean(s["dauern"]), 1)
                    ).number_format = "0.0"
    style_header(st, 5)

    # ---------------- Arbeitgeber (alle mit aktuell offenen Stellen)
    ag_stat = {}
    for r in log:
        s = ag_stat.setdefault(r["arbeitgeber"],
                               {"log": 0, "offen": 0, "dauern": []})
        s["log"] += 1
        if r["last_seen"] == stand:
            s["offen"] += 1
        else:
            s["dauern"].append(standzeit_tage(r))
    ag = wb.create_sheet("Arbeitgeber")
    ag.append(["Arbeitgeber", "Offene Stellen (aktuell)",
               "Stellen im Log gesamt", "Beendet",
               "Ø Standzeit beendeter Stellen (Tage)"])
    offene_ag = sorted((a for a, s in ag_stat.items() if s["offen"]),
                       key=lambda a: -ag_stat[a]["offen"])
    for i, a in enumerate(offene_ag, start=2):
        s = ag_stat[a]
        ag.cell(row=i, column=1, value=a)
        ag.cell(row=i, column=2, value=s["offen"])
        ag.cell(row=i, column=3, value=s["log"])
        ag.cell(row=i, column=4, value=len(s["dauern"]))
        if s["dauern"]:
            ag.cell(row=i, column=5,
                    value=round(statistics.mean(s["dauern"]), 1)
                    ).number_format = "0.0"
    style_header(ag, 5)

    # ---------------- Arbeitgeber-Tempo: wer besetzt am schnellsten?
    # Rangliste nach kuerzester mittlerer Standzeit beendeter Stellen.
    # Kurze Standzeit heisst nur "Anzeige schnell wieder offline" --
    # meist Besetzung, es kann aber auch Loeschung oder Ablauf sein.
    tempo = [(a, s) for a, s in ag_stat.items()
             if len(s["dauern"]) >= AG_TEMPO_MIN_BEENDET]
    tempo.sort(key=lambda t: (statistics.mean(t[1]["dauern"]),
                              -len(t[1]["dauern"])))
    tp = wb.create_sheet("Arbeitgeber-Tempo")
    tp.append(["Arbeitgeber", "Ø Standzeit (Tage)", "Median (Tage)",
               "Beendete Stellen", "Aktuell offen", "Stellen im Log",
               "Anteil beendet"])
    for i, (a, s) in enumerate(tempo, start=2):
        tp.cell(row=i, column=1, value=a)
        tp.cell(row=i, column=2,
                value=round(statistics.mean(s["dauern"]), 1)
                ).number_format = "0.0"
        tp.cell(row=i, column=3,
                value=round(statistics.median(s["dauern"]), 1)
                ).number_format = "0.0"
        tp.cell(row=i, column=4, value=len(s["dauern"]))
        tp.cell(row=i, column=5, value=s["offen"])
        tp.cell(row=i, column=6, value=s["log"])
        tp.cell(row=i, column=7,
                value=len(s["dauern"]) / s["log"]).number_format = "0%"
    style_header(tp, 7)
    hin = len(tempo) + 3
    tp.cell(row=hin, column=1, value=(
        f"Nur Arbeitgeber mit mindestens {AG_TEMPO_MIN_BEENDET} beendeten "
        "Stellen (Rauschen von Einzelfaellen). 'Anteil beendet' zeigt den "
        "Umschlag: hoch = es wird laufend besetzt und nachgelegt. Achtung "
        "bei der Deutung: Zeitarbeitsfirmen schalten viele kurzlaufende "
        "Anzeigen und stehen deshalb oft vorn; kurze Standzeit kann auch "
        "Loeschung oder Ablauf der Anzeige statt Besetzung sein."))
    if tempo:
        n_bar = min(20, len(tempo))
        bc = balken_chart(
            f"Schnellste Arbeitgeber: Ø Standzeit in Tagen "
            f"(min. {AG_TEMPO_MIN_BEENDET} beendete Stellen)",
            "Ø Tage", n_bar, breite=24)
        bc.add_data(Reference(tp, min_col=2, min_row=1, max_row=n_bar + 1),
                    titles_from_data=True)
        bc.set_categories(Reference(tp, min_col=1, min_row=2,
                                    max_row=n_bar + 1))
        bc.legend = None   # eine Serie, Legende ueberfluessig
        tp.add_chart(bc, f"A{hin + 2}")

    # ---------------- Arbeitslose (Kreisebene, Regionaldatenbank/GENESIS)
    al_daten = lade_arbeitslose()
    monate = sorted({(r["jahr"], r["monat"]) for r in al_daten})
    alm = {(r["jahr"], r["monat"], r["ags"]): r for r in al_daten}
    ags_liste = list(AL_KREISE)
    if al_daten:
        ab = wb.create_sheet("Arbeitslose")
        ab.append(["Monat"]
                  + [f"{AL_KREISE[a]} Arbeitslose" for a in ags_liste]
                  + [f"{AL_KREISE[a]} Quote %" for a in ags_liste])
        for i, (j, m) in enumerate(monate, start=2):
            ab.cell(row=i, column=1,
                    value=date(j, m, 1)).number_format = "MM.YYYY"
            for k, a in enumerate(ags_liste):
                r = alm.get((j, m, a), {})
                ab.cell(row=i, column=2 + k,
                        value=r.get("arbeitslose")).number_format = "#,##0"
                q = ab.cell(row=i, column=2 + len(ags_liste) + k,
                            value=r.get("quote"))
                q.number_format = "0.0"
        style_header(ab, 1 + 2 * len(ags_liste))
        nz = len(monate) + 1
        abc = linien_chart("Arbeitslose je Kreis (Monatswerte)",
                           "Monat", "Arbeitslose", "MM.YY", len(monate))
        abc.add_data(Reference(ab, min_col=2, max_col=1 + len(ags_liste),
                               min_row=1, max_row=nz),
                     titles_from_data=True)
        abc.set_categories(Reference(ab, min_col=1, min_row=2, max_row=nz))
        ab.add_chart(abc, f"A{nz + 3}")

    # ---------------- AL-Relation (Berufe manuell, Kreise automatisch)
    al = wb.create_sheet("AL-Relation")
    al.append(["Beruf", "Offene Stellen (aktuell)",
               "Arbeitslose (manuell, BA-Statistik)",
               "Arbeitslose je Stelle"])
    for i, b in enumerate(al_berufe, start=2):
        al.cell(row=i, column=1, value=b)
        # fester Wert statt Verweis auf Aktuell: so ist die Spalte auch
        # in Vorschau-Apps und auf dem Handy gefuellt
        al.cell(row=i, column=2, value=bges.get((stand, b), 0))
        c_in = al.cell(row=i, column=3)
        c_in.font = INPUT_FONT
        al.cell(row=i, column=4,
                value=f'=IF(OR(C{i}="",B{i}=0),"",ROUND(C{i}/B{i},1))')
        al.cell(row=i, column=4).number_format = "0.0"
    style_header(al, 4)

    start = len(al_berufe) + 4
    if al_daten:
        # Kreisblock unter der Berufstabelle: Arbeitslose (letzter Monat
        # aus GENESIS) je offener Stelle des passenden Suchgebiets
        j_l, m_l = monate[-1]
        al.cell(row=start, column=1,
                value=f"Kreisebene (Arbeitslose Stand {m_l:02d}/{j_l}, "
                      "Regionaldatenbank)")
        al.cell(row=start, column=1).font = Font(name=ARIAL, bold=True)
        for c, t in enumerate(["Kreis", "Arbeitslose",
                               "Offene Stellen (Suchgebiet)",
                               "Arbeitslose je Stelle"], start=1):
            al.cell(row=start + 1, column=c, value=t)
        style_header(al, 4, row=start + 1)
        zeile = start + 2
        # Teil-/Ueberlappungsgebiete hier auslassen: Waiblingen 30 km und
        # Backnang 10 km liegen im Rems-Murr-Kreis, Stuttgart Ost
        # (Fellbach 8 km) deckt den Stadtkreis nur minimal ab
        auslassen = {"waiblingen_30km", "backnang_10km", "stuttgart_ost"}
        for reg, ags in REGION_ZU_AGS.items():
            if reg not in aktiv or reg in auslassen or ags not in AL_KREISE:
                continue
            alo = alm.get((j_l, m_l, ags), {}).get("arbeitslose")
            stellen = snap_map.get((stand, reg))
            al.cell(row=zeile, column=1, value=AL_KREISE[ags])
            al.cell(row=zeile, column=2, value=alo).number_format = "#,##0"
            al.cell(row=zeile, column=3, value=stellen).number_format = "#,##0"
            if alo and stellen:
                al.cell(row=zeile, column=4,
                        value=round(alo / stellen, 1)).number_format = "0.0"
            zeile += 1
        al.cell(row=zeile + 1, column=1,
                value=("Hinweis: Suchgebiete sind Umkreissuchen und decken "
                       "die Kreise nur naeherungsweise ab. Arbeitslose je "
                       "Beruf gibt es nicht in der Regionaldatenbank, "
                       "die blaue Spalte oben bleibt daher manuell."))
    else:
        al.cell(row=start, column=1, value=(
            "Kreisdaten fehlen: data/arbeitslose.csv ist (noch) leer. "
            "GENESIS_USER/GENESIS_PASS als GitHub-Secrets setzen "
            "(kostenloses Konto auf regionalstatistik.de), dann fuellen "
            "sich das Blatt Arbeitslose und dieser Block automatisch."))

    # verbleibende Formeln (Relation zur blauen Spalte) beim Oeffnen
    # in Excel/LibreOffice sicher durchrechnen lassen
    wb.calculation.fullCalcOnLoad = True

    for ws in wb.worksheets:
        base_font(ws)
        autosize(ws)
    wb.save(OUT)
    print(f"Excel geschrieben: {OUT}")
    print(f"Stand {stand}, {len(daten)} Erhebungstage, "
          f"{len(log)} Stellen im Log.")
    return 0


if __name__ == "__main__":
    main()
    raise SystemExit(build_excel())
